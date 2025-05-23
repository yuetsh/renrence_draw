import random
import openpyxl
import pandas as pd


def draw_one(filepath, count):
    wb = openpyxl.load_workbook(filepath)
    course_count = len(wb.sheetnames)

    result = pd.DataFrame()

    for i in range(course_count):
        df = pd.read_excel(filepath, sheet_name=i, skiprows=1)
        df = pd.concat([df, result]).drop_duplicates("身份证号", keep=False)

        a = df[df["等级"] == "A"]
        b = df[df["等级"] == "B"]
        c = df[df["等级"] == "C"]

        random_a = a.sample(n=count)
        random_b = b.sample(n=count)
        random_c = c.sample(n=count)

        course = pd.concat([random_a, random_b, random_c])
        result = pd.concat([result, course])

    result["序号"] = range(1, 1 + len(result))
    return result


def draw_two(filepath, school1, school2):
    wb = openpyxl.load_workbook(filepath)
    course_count = len(wb.sheetnames)

    result = pd.DataFrame()

    for i in range(course_count):
        df = pd.read_excel(filepath, sheet_name=i, skiprows=1)
        df = pd.concat([df, result]).drop_duplicates("身份证号", keep=False)

        """
                学校 ———— 人数
        school1[0]  ———— school1[1]
        school2[0]  ———— school2[1]

        A8 B8 C8
        """

        """
        一个一个来
        """

        start = 1
        total = int((school1[1] + school2[1]) // 3)
        end = total - 1

        school1_rest = school1[1]
        school2_rest = school2[1]

        n1 = random.randint(start, end)
        a1 = df[(df["等级"] == "A") & (df["学校"] == school1[0])].sample(n=n1)
        school1_rest -= n1
        a2 = df[(df["等级"] == "A") & (df["学校"] == school2[0])].sample(n=(total - n1))
        school2_rest -= total - n1

        n2 = random.randint(start, end)
        b1 = df[(df["等级"] == "B") & (df["学校"] == school1[0])].sample(n=n2)
        school1_rest -= n2
        b2 = df[(df["等级"] == "B") & (df["学校"] == school2[0])].sample(n=(total - n2))
        school2_rest -= total - n2

        if school1_rest <= 0 or school2_rest <= 0:
            raise ValueError("抽取失败，请重新抽取")

        c1 = df[(df["等级"] == "C") & (df["学校"] == school1[0])].sample(n=school1_rest)
        c2 = df[(df["等级"] == "C") & (df["学校"] == school2[0])].sample(n=school2_rest)

        course = pd.concat([a1, a2, b1, b2, c1, c2])
        result = pd.concat([result, course])

    result["序号"] = range(1, 1 + len(result))
    return result


"""
users 是已经在语数英中抽到的学生 pf
filepath 是专业课文件路径
"""


def draw_zhuanyeke(users, filepath):
    wb = openpyxl.load_workbook(filepath)
    course_count = len(wb.sheetnames)

    result = pd.DataFrame()
    for i in range(course_count):
        df = pd.read_excel(filepath, sheet_name=i, skiprows=2)
        course = pd.DataFrame()
        if not df.empty:
            # 把 df 中出现在 users 里面的数据去掉
            df = df[~df["身份证号"].isin(users["身份证号"])]

            a = df[df["等级"] == "A"]
            b = df[df["等级"] == "B"]
            c = df[df["等级"] == "C"]

            n = 1 if i == 0 else 2

            random_a = a.sample(n=n)
            random_b = b.sample(n=n)
            random_c = c.sample(n=n)

            course = pd.concat([random_a, random_b, random_c])
        result = pd.concat([result, course])

    result["序号"] = range(1, 1 + len(result))
    return result
